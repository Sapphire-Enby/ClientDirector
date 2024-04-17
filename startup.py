#TODO:
# Header check returns properly now.
# Now we must give it the procedures for the checks it did
# if !file
#   makefile()
#       have this call fillfileHead()
# if emptyFile
#        fillfileHead()
#else
#   do LoadFile procedures
#   maybe just call a seperate file for handling loading of data
#   consider making the workbook and sheet global variables

import openpyxl
from pathlib import Path
#ofile = Path('./outfile.xlsx')
class Book:
    def __init__(self):
        self.wb = None
        self.s = None
        self.ofile = Path('./outfile.xlsx')
    
    #self management calls
    def load(self): self.wb = openpyxl.load_workbook(self.ofile)
    def save(self):self.wb.save(self.ofile)
    def act(self): self.s = self.wb.active
    def make(self):self.wb= openpyxl.Workbook() 
    def isempty(self):
        self.load()
        self.act()
        for row in self.s.iter_rows():
            for cell in row:
                if cell.value:
                    return False
                return True
    def writeHeader(self):
        self.act()
        for col, cellHeader in enumerate(["name", "address", "phone", "mail", "folder"],start=1):  # Start from column 1
            self.s.cell(row=1, column=col, value=cellHeader)

    def main(self):
    #self = Book()
        if not self.ofile.exists(): 
            self.make()
            self.act()
            self.writeHeader()
            self.save()

        elif self.isempty(): 
            self.writeHeader()
            self.save()
        else:
            pass
            #print("file Exists, and is not empty")
            #Proceed to load
        #self.save()        






def main():
    book = Book()
    if not book.ofile.exists(): 
        book.make()
        book.act()
        book.writeHeader()
        book.save()

    elif book.isempty(): 
        book.writeHeader()
        book.save()
    else:   
        print("file Exists, and is not empty")
        #Proceed to load
    book.save()
if __name__ == "__main__":
    main()
